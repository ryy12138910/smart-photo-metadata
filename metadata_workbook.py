#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Review-workbook generation and approved EXIF-writing helpers."""

from __future__ import print_function

from datetime import datetime
import hashlib
import os
import re
import shutil
import tempfile

from exif_utils import (
    copy_output_path,
    load_exif_dict,
    read_gps_exif,
    validate_lat_lon,
    write_datetime_exif,
    write_gps_exif,
    write_png_exif,
)


# Global defaults keep the program usable for projects in any region.
EXPECTED_LAT_MIN = -90.0
EXPECTED_LAT_MAX = 90.0
EXPECTED_LON_MIN = -180.0
EXPECTED_LON_MAX = 180.0

PHOTO_DISPLAY_MAX_WIDTH = 520
PHOTO_DISPLAY_MAX_HEIGHT = 320
THUMBNAIL_QUALITY = 70


# Chinese labels are stored as Unicode escapes to avoid encoding damage.
L_REVIEW = "\u5ba1\u6838\u7ed3\u679c"
L_PARSE_STATUS = "\u89e3\u6790\u72b6\u6001"
L_SOURCE = "\u5b57\u6bb5\u6765\u6e90"
L_OCR_TEXT = "OCR\u7ed3\u679c"
L_LON = "\u7ecf\u5ea6"
L_LAT = "\u7eac\u5ea6"
L_TIME = "\u62cd\u6444\u65f6\u95f4"
L_CONFIDENCE = "\u5b57\u6bb5\u7f6e\u4fe1\u5ea6"
L_PHOTO = "\u7167\u7247"
L_PHOTO_LINK = "\u7167\u7247\u94fe\u63a5"
L_SOURCE_PATH = "\u539f\u56fe\u8def\u5f84"
L_TIPS = "\u89e3\u6790\u63d0\u793a"
L_MODEL_NOTE = "\u6a21\u578b\u8bf4\u660e"
L_NOTE = "\u5ba1\u6838\u5907\u6ce8"

V_UNREVIEWED = "\u672a\u5ba1"
V_APPROVED = "\u901a\u8fc7"
V_AUTO_APPROVED = "\u81ea\u52a8\u901a\u8fc7"
V_CHECK = "\u5f85\u6838"
V_REJECTED = "\u4e0d\u901a\u8fc7"
V_SKIP = "\u8df3\u8fc7"
V_YES = "\u662f"
STATUS_EXIF_COMPLETE = "EXIF\u5b8c\u6574"

APPROVED_VALUES = (V_APPROVED, V_AUTO_APPROVED, V_YES, "YES", "Y", "OK", "1", "TRUE")
REVIEW_HEADERS = [
    L_REVIEW,
    L_PARSE_STATUS,
    L_SOURCE,
    L_OCR_TEXT,
    L_LON,
    L_LAT,
    L_TIME,
    L_CONFIDENCE,
    L_PHOTO,
    L_PHOTO_LINK,
    L_SOURCE_PATH,
    L_TIPS,
    L_MODEL_NOTE,
    L_NOTE,
]

EN_HEADER_BY_ZH = {
    L_REVIEW: "Review Decision",
    L_PARSE_STATUS: "Parse Status",
    L_SOURCE: "Field Source",
    L_OCR_TEXT: "OCR Result",
    L_LON: "Longitude",
    L_LAT: "Latitude",
    L_TIME: "Capture Time",
    L_CONFIDENCE: "Field Confidence",
    L_PHOTO: "Photo",
    L_PHOTO_LINK: "Photo Link",
    L_SOURCE_PATH: "Source Path",
    L_TIPS: "Review Tips",
    L_MODEL_NOTE: "Model Notes",
    L_NOTE: "Reviewer Notes",
}
ZH_HEADER_BY_EN = {value: key for key, value in EN_HEADER_BY_ZH.items()}
EN_DECISION_BY_ZH = {
    V_UNREVIEWED: "Unreviewed",
    V_APPROVED: "Approved",
    V_AUTO_APPROVED: "Auto Approved",
    V_CHECK: "Needs Review",
    V_REJECTED: "Rejected",
    V_SKIP: "Skip",
}
ZH_DECISION_BY_EN = {value: key for key, value in EN_DECISION_BY_ZH.items()}
APPROVED_VALUES = APPROVED_VALUES + ("APPROVED", "AUTO APPROVED")

IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".bmp", ".webp")
JPG_EXTENSIONS = (".jpg", ".jpeg")
WRITE_EXIF_EXTENSIONS = (".jpg", ".jpeg", ".png")
SKIP_EMBED_IMAGE_EXTENSIONS = (".mpo",)
SKIP_EMBED_IMAGE_FORMATS = ("MPO",)
PROJECT_LAT_MIN = 30.0
PROJECT_LAT_MAX = 33.0
PROJECT_LON_MIN = 119.0
PROJECT_LON_MAX = 122.0


def clean_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_path(value):
    return clean_cell(value).replace("/", os.sep).replace("\\", os.sep)


def normalize_text(text):
    text = clean_cell(text)
    replacements = {
        "\uff0c": ",",
        "\uff1a": ":",
        "\uff1b": ";",
        "\u3002": ".",
        "\u2032": "'",
        "\u2019": "'",
        "\u2018": "'",
        "\u2033": '"',
        "\u201c": '"',
        "\u201d": '"',
        "\u00b0": "\u5ea6",
        "\u00ba": "\u5ea6",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    text = re.sub(r'"{2,}', '"', text)
    text = re.sub(r"'{2,}", "'", text)
    return re.sub(r"\s+", " ", text).strip()


class ParseWarning(ValueError):
    pass


def dms_to_decimal(degree, minute, second, kind):
    degree = int(float(degree))
    minute = int(float(minute))
    second = int(float(second))
    min_degree = 0 if kind == "lat" else 70
    max_degree = 90 if kind == "lat" else 180

    if not (min_degree <= degree <= max_degree):
        raise ParseWarning("%s degree out of range: %s" % (kind, degree))
    if not (0 <= minute < 60):
        raise ParseWarning("%s minute out of range: %s" % (kind, minute))
    if not (0 <= second < 60):
        raise ParseWarning("%s second out of range: %s" % (kind, second))
    return degree + minute / 60.0 + second / 3600.0


def dms_display(degree, minute, second=0):
    return "%d\u5ea6%02d\u5206%02d\u79d2" % (int(float(degree)), int(float(minute)), int(float(second)))


def split_minute_second(minute_text, second_text=""):
    minute_text = clean_cell(minute_text)
    second_text = clean_cell(second_text)
    if second_text:
        return minute_text, second_text
    if minute_text.isdigit() and len(minute_text) > 2:
        return minute_text[:-2], minute_text[-2:]
    return minute_text, "0"


def compact_decimal_from_digits(digits, kind):
    deg_digits = 2 if kind == "lat" else 3
    min_degree = 0 if kind == "lat" else 70
    max_degree = 90 if kind == "lat" else 180
    if len(digits) <= deg_digits + 1:
        raise ParseWarning("%s compact decimal is too short: %s" % (kind, digits))
    display = "%s.%s" % (digits[:deg_digits], digits[deg_digits:])
    value = float(display)
    if min_degree <= abs(value) <= max_degree:
        return value, display
    raise ParseWarning("%s compact decimal out of range: %s" % (kind, display))


def project_coordinate_in_range(value, kind):
    if kind == "lat":
        return PROJECT_LAT_MIN <= abs(value) <= PROJECT_LAT_MAX
    return PROJECT_LON_MIN <= abs(value) <= PROJECT_LON_MAX


def expected_coordinate_in_range(value, kind):
    if kind == "lat":
        return EXPECTED_LAT_MIN <= abs(value) <= EXPECTED_LAT_MAX
    return EXPECTED_LON_MIN <= abs(value) <= EXPECTED_LON_MAX


def likely_compact_date_digits(digits):
    if len(digits) != 8 or not digits.startswith(("19", "20")):
        return False
    try:
        month = int(digits[4:6])
        day = int(digits[6:8])
    except ValueError:
        return False
    return 1 <= month <= 12 and 1 <= day <= 31


def decimal_match_is_embedded_in_date(text, start, end):
    before = text[start - 1] if start > 0 else ""
    before_before = text[start - 2] if start > 1 else ""
    after = text[end] if end < len(text) else ""
    after_after = text[end + 1] if end + 1 < len(text) else ""
    if before == "." and before_before.isdigit():
        return True
    if after == "." and after_after.isdigit():
        return True
    return False


def compact_decimal_candidate(segment, kind, project_only=False):
    deg_digits = 2 if kind == "lat" else 3
    min_len = 4 if kind == "lat" else 5
    max_len = deg_digits + 6
    for match in re.finditer(r"(?<!\d)(\d{%d,%d})(?!\d)" % (min_len, max_len), segment):
        digits = match.group(1)
        if likely_compact_date_digits(digits):
            continue
        try:
            value, display = compact_decimal_from_digits(digits, kind)
        except ParseWarning:
            continue

        if len(digits) == deg_digits + 2:
            degree = digits[:deg_digits]
            minute = digits[deg_digits:]
            try:
                dms_value = dms_to_decimal(degree, minute, 0, kind)
                dms_text = dms_display(degree, minute, 0)
            except ParseWarning:
                dms_value = None
                dms_text = ""
            if dms_value is not None:
                decimal_expected = expected_coordinate_in_range(value, kind)
                dms_expected = expected_coordinate_in_range(dms_value, kind)
                if dms_expected and not decimal_expected:
                    return dms_value, dms_text
                if decimal_expected:
                    return value, display

        if project_only and not project_coordinate_in_range(value, kind):
            continue
        return value, display
    raise ValueError("No compact decimal %s candidate found" % kind)


def first_segment_after(text, labels, stop_labels):
    best = None
    best_label = ""
    for label in labels:
        match = re.search(re.escape(label), text, re.I)
        if match and (best is None or match.start() < best.start()):
            best = match
            best_label = match.group(0)
    if not best:
        return "", ""

    start = best.end()
    end = len(text)
    for label in stop_labels:
        match = re.search(re.escape(label), text[start:], re.I)
        if match:
            end = min(end, start + match.start())
    return text[start:end], best_label


def parse_coordinate_segment(segment, kind):
    segment = normalize_text(segment)
    min_degree = 0 if kind == "lat" else 70
    max_degree = 90 if kind == "lat" else 180

    decimal = None
    for match in re.finditer(r"(?<!\d)[-+]?\d{1,3}(?:\.\d+|:\d{3,8})(?!\d)", segment):
        if decimal_match_is_embedded_in_date(segment, match.start(), match.end()):
            continue
        decimal = match
        break
    if decimal:
        value = float(decimal.group(0).replace(":", "."))
        if min_degree <= abs(value) <= max_degree:
            return value
        raise ParseWarning("%s decimal coordinate out of range: %s" % (kind, value))

    try:
        value, _ = compact_decimal_candidate(segment, kind)
        return value
    except Exception:
        pass

    glued_dms = re.search(
        r"(\d{1,3})\s*(?:\u5ea6|deg|d)\s*(\d{3,4})(?:\s*(?:\u79d2|\"+))?",
        segment,
        re.I,
    )
    if glued_dms:
        minute, second = split_minute_second(glued_dms.group(2))
        return dms_to_decimal(glued_dms.group(1), minute, second, kind)

    degree_minute = re.search(
        r"(\d{1,3})\s*(?:\u5ea6|deg|d)\s*(\d{1,2})(?!\d|\s*(?:\u5206|'|\u79d2|\"))",
        segment,
        re.I,
    )
    if degree_minute:
        return dms_to_decimal(degree_minute.group(1), degree_minute.group(2), 0, kind)

    numbers = re.findall(r"\d+(?:\.\d+)?", segment)
    if len(numbers) >= 3:
        return dms_to_decimal(numbers[0], numbers[1], numbers[2], kind)
    if len(numbers) >= 2:
        return dms_to_decimal(numbers[0], numbers[1], 0, kind)

    raise ValueError("Could not parse %s coordinate from %r" % (kind, segment))


def coordinate_display_from_segment(segment, kind):
    segment = normalize_text(segment)
    decimal = None
    for match in re.finditer(r"(?<!\d)([-+]?\d{1,3}(?:\.\d+|:\d{3,8}))(?!\d)", segment):
        if decimal_match_is_embedded_in_date(segment, match.start(), match.end()):
            continue
        decimal = match
        break
    if decimal:
        return decimal.group(1).replace(":", ".")

    try:
        _, display = compact_decimal_candidate(segment, kind)
        return display
    except Exception:
        pass

    glued_dms = re.search(
        r"(\d{1,3})\s*(?:\u5ea6|deg|d)\s*(\d{3,4})(?:\s*(?:\u79d2|\"+))?",
        segment,
        re.I,
    )
    if glued_dms:
        minute, second = split_minute_second(glued_dms.group(2))
        return dms_display(glued_dms.group(1), minute, second)

    degree_minute = re.search(
        r"(\d{1,3})\s*(?:\u5ea6|deg|d)\s*(\d{1,2})(?!\d|\s*(?:\u5206|'|\u79d2|\"))",
        segment,
        re.I,
    )
    if degree_minute:
        return dms_display(degree_minute.group(1), degree_minute.group(2), 0)

    numbers = re.findall(r"\d+(?:\.\d+)?", segment)
    if len(numbers) >= 3:
        return dms_display(numbers[0], numbers[1], numbers[2])
    if len(numbers) >= 2:
        return dms_display(numbers[0], numbers[1], 0)
    return ""


def directed_decimal_pair(text):
    clean = normalize_text(text)
    pattern = r"([NSEW\u5317\u5357\u4e1c\u897f])?\s*([-+]?\d{1,3}\.\d+)\s*([NSEW\u5317\u5357\u4e1c\u897f])?"
    lat_values = []
    lon_values = []
    for match in re.finditer(pattern, clean, re.I):
        ref = clean_cell(match.group(1) or match.group(3)).upper()
        if not ref:
            continue
        value = abs(float(match.group(2)))
        display = match.group(2)
        if ref in ("N", "S", "\u5317", "\u5357") and value <= 90:
            lat_values.append((value, display))
        elif ref in ("E", "W", "\u4e1c", "\u897f") and value <= 180:
            lon_values.append((value, display))
    if lat_values and lon_values:
        return lat_values[0], lon_values[0]
    raise ValueError("No directed decimal pair found")


def directed_dms_pair(text):
    clean = normalize_text(text)
    prefix_pattern = (
        r"(^|[\s,;])([NSEW\u5317\u5357\u4e1c\u897f])\s*"
        r"(\d{1,3}\s*(?:\u5ea6|deg|d)?\s*"
        r"\d{1,2}(?:\s*(?:\u5206|'))?\s*"
        r"(?:\d{1,2}\s*(?:\u79d2|\"+))?)"
    )
    clean = re.sub(prefix_pattern, r"\1\3\2", clean, flags=re.I)
    lat_values = []
    lon_values = []

    def add_candidate(ref, degree, minute_text, second_text):
        ref = clean_cell(ref).upper()
        if not ref:
            return
        minute, second = split_minute_second(minute_text, second_text)
        if ref in ("N", "S", "\u5317", "\u5357"):
            try:
                value = dms_to_decimal(degree, minute, second, "lat")
            except ParseWarning:
                return
            lat_values.append((abs(value), dms_display(degree, minute, second)))
        elif ref in ("E", "W", "\u4e1c", "\u897f"):
            try:
                value = dms_to_decimal(degree, minute, second, "lon")
            except ParseWarning:
                return
            lon_values.append((abs(value), dms_display(degree, minute, second)))

    glued_pattern = (
        r"([NSEW\u5317\u5357\u4e1c\u897f])?\s*"
        r"(\d{1,3})\s*(?:\u5ea6|deg|d)\s*"
        r"(\d{3,4})\s*(?:\u79d2|\"+)?\s*"
        r"([NSEW\u5317\u5357\u4e1c\u897f])?"
    )
    for match in re.finditer(glued_pattern, clean, re.I):
        add_candidate(match.group(4) or match.group(1), match.group(2), match.group(3), "")

    pattern = (
        r"([NSEW\u5317\u5357\u4e1c\u897f])?\s*"
        r"(\d{1,3})\s*(?:\u5ea6|deg|d)?\s*"
        r"(\d{1,2})(?:\s*(?:\u5206|'))?\s*"
        r"(?:(\d{1,2})\s*(?:\u79d2|\"+))?\s*"
        r"([NSEW\u5317\u5357\u4e1c\u897f])?"
    )
    for match in re.finditer(pattern, clean, re.I):
        add_candidate(match.group(5) or match.group(1), match.group(2), match.group(3), match.group(4))

    if lat_values and lon_values:
        return lat_values[0], lon_values[0]
    raise ValueError("No directed DMS pair found")


LAT_LABELS = ["\u5317\u7eac", "\u5357\u7eac", "\u7eac\u5ea6", "\u7eac", "GPSLatitude", "Latitude", "Lat"]
LON_LABELS = ["\u4e1c\u7ecf", "\u897f\u7ecf", "\u7ecf\u5ea6", "\u7ecf", "GPSLongitude", "Longitude", "Lng", "Lon", "Long"]


def first_labeled_coordinate(text, labels, kind):
    matches = []
    for label in labels:
        for match in re.finditer(re.escape(label), text, re.I):
            matches.append(match)
    matches.sort(key=lambda item: item.start())
    for match in matches:
        segment = text[match.end():]
        try:
            value = parse_coordinate_segment(segment, kind)
            display = coordinate_display_from_segment(segment, kind)
        except Exception:
            continue
        if display and project_coordinate_in_range(value, kind):
            return display, abs(value)
    return None


def labeled_coordinate_values(text):
    clean = normalize_text(text)
    return (
        first_labeled_coordinate(clean, LAT_LABELS, "lat"),
        first_labeled_coordinate(clean, LON_LABELS, "lon"),
    )


def infer_unlabeled_coordinates(text):
    clean = normalize_text(text)
    lat_pair = None
    lon_pair = None

    for match in re.finditer(r"(?<!\d)([-+]?\d{1,3}(?:\.\d+|:\d{3,8}))(?!\d)", clean):
        if decimal_match_is_embedded_in_date(clean, match.start(), match.end()):
            continue
        display = match.group(1).replace(":", ".")
        value = abs(float(display))
        if project_coordinate_in_range(value, "lon") and lon_pair is None:
            lon_pair = (display, value)
        elif project_coordinate_in_range(value, "lat") and lat_pair is None:
            lat_pair = (display, value)
        if lat_pair and lon_pair:
            return lat_pair, lon_pair

    try:
        value, display = compact_decimal_candidate(clean, "lat", project_only=True)
        lat_pair = (display, abs(value))
    except Exception:
        pass
    try:
        value, display = compact_decimal_candidate(clean, "lon", project_only=True)
        lon_pair = (display, abs(value))
    except Exception:
        pass
    if lat_pair and lon_pair:
        return lat_pair, lon_pair

    pattern = (
        r"(?<!\d)(\d{1,3})\s*(?:\u5ea6|deg|d)\s*"
        r"(\d{1,2})(?:\s*(?:\u5206|'))?\s*"
        r"(?:(\d{1,2})\s*(?:\u79d2|\"+))?"
    )
    for match in re.finditer(pattern, clean, re.I):
        degree = match.group(1)
        minute = match.group(2)
        second = match.group(3) or "0"
        degree_value = int(degree)
        if 70 <= degree_value <= 180:
            try:
                value = dms_to_decimal(degree, minute, second, "lon")
            except ParseWarning:
                continue
            if project_coordinate_in_range(value, "lon") and lon_pair is None:
                lon_pair = (dms_display(degree, minute, second), abs(value))
        if 0 <= degree_value <= 90:
            try:
                value = dms_to_decimal(degree, minute, second, "lat")
            except ParseWarning:
                continue
            if project_coordinate_in_range(value, "lat") and lat_pair is None:
                lat_pair = (dms_display(degree, minute, second), abs(value))
        if lat_pair and lon_pair:
            break
    return lat_pair, lon_pair


def gps_candidates_from_ocr(text):
    clean = normalize_text(text)
    for parser in (directed_decimal_pair, directed_dms_pair):
        try:
            lat_pair, lon_pair = parser(clean)
            return lat_pair[1], lon_pair[1], lat_pair[0], lon_pair[0]
        except Exception:
            pass

    lat_pair, lon_pair = labeled_coordinate_values(clean)
    if (lat_pair and not lon_pair) or (lon_pair and not lat_pair):
        inferred_lat, inferred_lon = infer_unlabeled_coordinates(clean)
        lat_pair = lat_pair or inferred_lat
        lon_pair = lon_pair or inferred_lon
    if lat_pair or lon_pair:
        return (
            lat_pair[0] if lat_pair else "",
            lon_pair[0] if lon_pair else "",
            lat_pair[1] if lat_pair else None,
            lon_pair[1] if lon_pair else None,
        )

    lat_pair, lon_pair = infer_unlabeled_coordinates(clean)
    if lat_pair or lon_pair:
        return (
            lat_pair[0] if lat_pair else "",
            lon_pair[0] if lon_pair else "",
            lat_pair[1] if lat_pair else None,
            lon_pair[1] if lon_pair else None,
        )

    lat, lon = parse_gps_from_ocr(clean)
    return "%.8f" % lat, "%.8f" % lon, lat, lon


def parse_gps_from_ocr(text):
    clean = normalize_text(text)
    try:
        lat_pair, lon_pair = directed_decimal_pair(clean)
        return lat_pair[0], lon_pair[0]
    except Exception:
        pass
    try:
        lat_pair, lon_pair = directed_dms_pair(clean)
        return lat_pair[0], lon_pair[0]
    except Exception:
        pass

    lat_pair, lon_pair = labeled_coordinate_values(clean)
    if lat_pair and lon_pair:
        return lat_pair[1], lon_pair[1]

    lat_labels = LAT_LABELS
    lon_labels = LON_LABELS
    lat_segment, lat_label = first_segment_after(clean, lat_labels, lon_labels)
    lon_segment, lon_label = first_segment_after(clean, lon_labels, [])

    if lat_segment and lon_segment:
        lat = parse_coordinate_segment(lat_segment, "lat")
        lon = parse_coordinate_segment(lon_segment, "lon")
        return abs(lat), abs(lon)

    numbers = [float(v) for v in re.findall(r"[-+]?\d{1,3}\.\d+", clean)]
    for index in range(0, len(numbers) - 1):
        lat = numbers[index]
        lon = numbers[index + 1]
        if -90 <= lat <= 90 and 70 <= abs(lon) <= 180:
            return abs(lat), abs(lon)
    raise ValueError("No latitude/longitude pair found in OCR text")


def gps_display_from_ocr(text):
    clean = normalize_text(text)
    try:
        lat_pair, lon_pair = directed_decimal_pair(clean)
        return lat_pair[1], lon_pair[1], lat_pair[0], lon_pair[0]
    except Exception:
        pass
    try:
        lat_pair, lon_pair = directed_dms_pair(clean)
        return lat_pair[1], lon_pair[1], lat_pair[0], lon_pair[0]
    except Exception:
        pass

    lat_pair, lon_pair = labeled_coordinate_values(clean)
    if lat_pair and lon_pair:
        return lat_pair[0], lon_pair[0], lat_pair[1], lon_pair[1]

    lat_labels = LAT_LABELS
    lon_labels = LON_LABELS
    lat_segment, lat_label = first_segment_after(clean, lat_labels, lon_labels)
    lon_segment, lon_label = first_segment_after(clean, lon_labels, [])
    if lat_segment and lon_segment:
        lat = parse_coordinate_segment(lat_segment, "lat")
        lon = parse_coordinate_segment(lon_segment, "lon")
        lat_display = coordinate_display_from_segment(lat_segment, "lat")
        lon_display = coordinate_display_from_segment(lon_segment, "lon")
        return lat_display, lon_display, abs(lat), abs(lon)

    lat, lon = parse_gps_from_ocr(clean)
    return "%.8f" % lat, "%.8f" % lon, lat, lon


def build_datetime(year, month, day, hour=0, minute=0, second=0):
    hour = int(hour or 0)
    minute = int(minute or 0)
    second = int(second or 0)
    if not (0 <= hour <= 23 and 0 <= minute <= 59 and 0 <= second <= 59):
        raise ValueError("time out of range")
    return datetime(int(year), int(month), int(day), hour, minute, second)


def datetime_from_match(match):
    return build_datetime(
        match.group(1),
        match.group(2),
        match.group(3),
        match.group(4) or 0,
        match.group(5) or 0,
        match.group(6) or 0,
    )


def first_date_candidate(text):
    date_patterns = [
        r"(?<!\d)((?:19|20)\d{2})[-:/\.](\d{1,2})[-:/\.](\d{1,2})(?!\d)",
        r"(?<!\d)((?:19|20)\d{2})\u5e74\s*(\d{1,2})\u6708\s*(\d{1,2})\u65e5?",
        r"(?<!\d)((?:19|20)\d{2})(\d{2})(\d{2})(?!\d)",
    ]
    for pattern in date_patterns:
        for match in re.finditer(pattern, text):
            try:
                build_datetime(match.group(1), match.group(2), match.group(3))
            except ValueError:
                continue
            return match.group(1), match.group(2), match.group(3)
    return None


def first_time_candidate(text):
    for match in re.finditer(r"(?<!\d)([01]?\d|2[0-3])[:\uff1a\.]([0-5]\d)(?:[:\uff1a\.]([0-5]\d))?(?!\d)", text):
        if decimal_match_is_embedded_in_date(text, match.start(), match.end()):
            continue
        return match.group(1), match.group(2), match.group(3) or 0
    return None


def global_datetime_candidate(text):
    date_parts = first_date_candidate(text)
    time_parts = first_time_candidate(text)
    if not date_parts or not time_parts:
        return None
    return build_datetime(
        date_parts[0],
        date_parts[1],
        date_parts[2],
        time_parts[0],
        time_parts[1],
        time_parts[2],
    )


def parse_datetime_from_ocr(text):
    clean = normalize_text(text)
    try:
        value = global_datetime_candidate(clean)
        if value:
            return value.strftime("%Y:%m:%d %H:%M:%S")
    except ValueError:
        pass

    time_before_date_patterns = [
        r"(?<!\d)([01]?\d|2[0-3])[:\uff1a\.]([0-5]\d)(?:[:\uff1a\.]([0-5]\d))?[^\d]{0,24}((?:19|20)\d{2})[-:/\.](\d{1,2})[-:/\.](\d{1,2})",
        r"(?<!\d)([01]?\d|2[0-3])[:\uff1a\.]([0-5]\d)(?:[:\uff1a\.]([0-5]\d))?[^\d]{0,24}((?:19|20)\d{2})\u5e74\s*(\d{1,2})\u6708\s*(\d{1,2})\u65e5?",
    ]
    for pattern in time_before_date_patterns:
        for match in re.finditer(pattern, clean):
            try:
                value = build_datetime(
                    match.group(4),
                    match.group(5),
                    match.group(6),
                    match.group(1),
                    match.group(2),
                    match.group(3) or 0,
                )
            except ValueError:
                continue
            return value.strftime("%Y:%m:%d %H:%M:%S")

    loose_patterns = [
        r"(?<!\d)((?:19|20)\d{2})(\d{2})[-:/\.](\d{1,2})(?:[^\d]{0,24}([01]?\d|2[0-3])[:\uff1a\.]([0-5]\d)(?:[:\uff1a\.]([0-5]\d))?)?",
        r"(?<!\d)((?:19|20)\d{2})(\d{2})\u6708\s*(\d{1,2})\u65e5?(?:[^\d]{0,24}([01]?\d|2[0-3])[:\uff1a\.]([0-5]\d)(?:[:\uff1a\.]([0-5]\d))?)?",
    ]
    for pattern in loose_patterns:
        for match in re.finditer(pattern, clean):
            try:
                value = datetime_from_match(match)
            except ValueError:
                continue
            return value.strftime("%Y:%m:%d %H:%M:%S")

    meridiem_patterns = [
        r"(?<!\d)((?:19|20)\d{2})[-:/\.](\d{1,2})[-:/\.](\d{1,2})[^\d]{0,12}(\u4e0a\u5348|\u4e0b\u5348|\u4e2d\u5348|\u665a\u4e0a|\u51cc\u6668|AM|PM|A\.M\.|P\.M\.)\s*(\d{1,2})(?:[:\uff1a\.\u70b9\u65f6](\d{1,2}))?(?:[:\uff1a\u5206](\d{1,2}))?",
        r"(?<!\d)((?:19|20)\d{2})\u5e74\s*(\d{1,2})\u6708\s*(\d{1,2})\u65e5?[^\d]{0,12}(\u4e0a\u5348|\u4e0b\u5348|\u4e2d\u5348|\u665a\u4e0a|\u51cc\u6668|AM|PM|A\.M\.|P\.M\.)\s*(\d{1,2})(?:[:\uff1a\.\u70b9\u65f6](\d{1,2}))?(?:[:\uff1a\u5206](\d{1,2}))?",
    ]
    for pattern in meridiem_patterns:
        for match in re.finditer(pattern, clean, re.I):
            marker = match.group(4).upper()
            hour = int(match.group(5) or 0)
            minute = int(match.group(6) or 0)
            second = int(match.group(7) or 0)
            if marker in ("\u4e0b\u5348", "\u665a\u4e0a", "\u4e2d\u5348", "PM", "P.M.") and hour < 12:
                hour += 12
            elif marker in ("\u4e0a\u5348", "\u51cc\u6668", "AM", "A.M.") and hour == 12:
                hour = 0
            try:
                value = datetime(
                    int(match.group(1)),
                    int(match.group(2)),
                    int(match.group(3)),
                    hour,
                    minute,
                    second,
                )
            except ValueError:
                continue
            return value.strftime("%Y:%m:%d %H:%M:%S")

    patterns = [
        r"(?<!\d)((?:19|20)\d{2})[-:/\.](\d{1,2})[-:/\.](\d{1,2})(?:[^\d]{0,8}([01]?\d|2[0-3])[:\uff1a\.]([0-5]\d)(?:[:\uff1a\.]([0-5]\d))?)?",
        r"(?<!\d)((?:19|20)\d{2})\u5e74\s*(\d{1,2})\u6708\s*(\d{1,2})\u65e5?(?:[^\d]{0,8}([01]?\d|2[0-3])[\u65f6:\uff1a\.]([0-5]\d)(?:[\u5206:\uff1a\.]([0-5]\d))?)?",
        r"(?<!\d)((?:19|20)\d{2})(\d{2})(\d{2})(?:[^\d]{1,8}(\d{2})(\d{2})(?:([0-5]\d))?)?(?!\d)",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, clean):
            try:
                value = datetime_from_match(match)
            except ValueError:
                continue
            return value.strftime("%Y:%m:%d %H:%M:%S")
    return ""


def add_original_image_to_sheet(sheet, image_path, cell, max_width, max_height):
    from openpyxl.drawing.image import Image as ExcelImage
    from PIL import Image as PILImage

    if os.path.splitext(image_path)[1].lower() in SKIP_EMBED_IMAGE_EXTENSIONS:
        return False

    with PILImage.open(image_path) as img:
        image_format = (img.format or "").upper()
        width, height = img.size
    if image_format in SKIP_EMBED_IMAGE_FORMATS:
        return False
    if width <= 0 or height <= 0:
        return False

    excel_image = ExcelImage(image_path)
    scale = min(float(max_width) / float(width), float(max_height) / float(height), 1.0)
    excel_image.width = int(width * scale)
    excel_image.height = int(height * scale)
    sheet.add_image(excel_image, cell)
    return True


def create_excel_thumbnail(
    image_path,
    temp_dir,
    max_width,
    max_height,
    quality=70,
):
    """Create a small JPEG preview suitable for embedding in an XLSX."""
    from PIL import Image as PILImage
    from PIL import ImageOps

    quality = max(30, min(95, int(quality)))
    digest = hashlib.sha1(os.path.abspath(image_path).encode("utf-8")).hexdigest()
    thumbnail_path = os.path.join(temp_dir, digest + ".jpg")
    with PILImage.open(image_path) as source:
        image = ImageOps.exif_transpose(source)
        if image.mode in ("RGBA", "LA") or (
            image.mode == "P" and "transparency" in image.info
        ):
            rgba = image.convert("RGBA")
            background = PILImage.new("RGB", rgba.size, "white")
            background.paste(rgba, mask=rgba.getchannel("A"))
            image = background
        elif image.mode != "RGB":
            image = image.convert("RGB")
        image.thumbnail(
            (max(1, int(max_width)), max(1, int(max_height))),
            PILImage.Resampling.LANCZOS,
        )
        image.save(
            thumbnail_path,
            format="JPEG",
            quality=quality,
            optimize=True,
        )
    return thumbnail_path


def add_review_image_to_sheet(
    sheet,
    image_path,
    cell,
    max_width,
    max_height,
    image_mode,
    temp_dir,
    thumbnail_quality,
):
    if image_mode == "none":
        return False
    if image_mode == "original":
        return add_original_image_to_sheet(
            sheet,
            image_path,
            cell,
            max_width,
            max_height,
        )

    from openpyxl.drawing.image import Image as ExcelImage

    thumbnail_path = create_excel_thumbnail(
        image_path,
        temp_dir,
        max_width,
        max_height,
        thumbnail_quality,
    )
    excel_image = ExcelImage(thumbnail_path)
    sheet.add_image(excel_image, cell)
    return True


def write_review_workbook(
    xlsx_path,
    records,
    max_width,
    max_height,
    image_mode="thumbnail",
    thumbnail_quality=70,
    language="zh",
):
    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        raise RuntimeError("Missing dependency: openpyxl. Install it with: pip install openpyxl")

    out_dir = os.path.dirname(os.path.abspath(xlsx_path))
    if out_dir and not os.path.isdir(out_dir):
        os.makedirs(out_dir)

    if image_mode not in ("thumbnail", "none", "original"):
        raise ValueError("unsupported Excel image mode: %s" % image_mode)

    with tempfile.TemporaryDirectory(prefix="metadata_excel_thumbs_") as temp_dir:
        workbook = Workbook()
        sheet = workbook.active
        english = language == "en"
        output_headers = [EN_HEADER_BY_ZH[header] for header in REVIEW_HEADERS] if english else REVIEW_HEADERS
        sheet.title = "OCR Review" if english else "OCR\u5ba1\u6838"
        sheet.append(output_headers)

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        status_fills = {
            "OK": PatternFill("solid", fgColor="DFF0D8"),
            "AUTO_OK": PatternFill("solid", fgColor="C6EFCE"),
            "NEEDS_REVIEW": PatternFill("solid", fgColor="FFF2CC"),
            "SUSPICIOUS": PatternFill("solid", fgColor="FCF8E3"),
            "FAIL": PatternFill("solid", fgColor="F2DEDE"),
            STATUS_EXIF_COMPLETE: PatternFill("solid", fgColor="D9EAD3"),
        }

        decision_values = [
            V_UNREVIEWED, V_APPROVED, V_AUTO_APPROVED, V_CHECK, V_REJECTED, V_SKIP
        ]
        if english:
            decision_values = [EN_DECISION_BY_ZH[value] for value in decision_values]
        validation = DataValidation(
            type="list",
            formula1='"%s"' % ",".join(decision_values),
            allow_blank=False,
        )
        sheet.add_data_validation(validation)
        header_indexes = {header: index + 1 for index, header in enumerate(REVIEW_HEADERS)}
        photo_col = header_indexes[L_PHOTO]
        link_col = header_indexes[L_PHOTO_LINK]

        for index, record in enumerate(records, start=1):
            values = [record.get(header, "") for header in REVIEW_HEADERS]
            if english:
                values[0] = EN_DECISION_BY_ZH.get(values[0], values[0])
                if values[1] == STATUS_EXIF_COMPLETE:
                    values[1] = "EXIF Complete"
            sheet.append(values)
            row_index = sheet.max_row
            if image_mode == "none":
                sheet.row_dimensions[row_index].height = 45
            else:
                sheet.row_dimensions[row_index].height = max(120, int(max_height * 0.78))
            validation.add(sheet.cell(row_index, 1))

            fill = status_fills.get(record.get(L_PARSE_STATUS))
            if fill:
                sheet.cell(row_index, 2).fill = fill

            image_path = record.get(L_SOURCE_PATH, "")
            if image_path and os.path.isfile(image_path):
                try:
                    from openpyxl.utils import get_column_letter

                    embedded = add_review_image_to_sheet(
                        sheet,
                        image_path,
                        "%s%d" % (get_column_letter(photo_col), row_index),
                        max_width,
                        max_height,
                        image_mode,
                        temp_dir,
                        thumbnail_quality,
                    )
                    if not embedded:
                        sheet.cell(row_index, photo_col).value = (
                            "Not embedded; use the photo link"
                            if english
                            else "\u672a\u5d4c\u5165\uff1b\u8bf7\u4f7f\u7528\u7167\u7247\u94fe\u63a5"
                        )
                except Exception as exc:
                    sheet.cell(row_index, photo_col).value = "image failed: %s" % exc

            link_cell = sheet.cell(row_index, link_col)
            if image_path:
                link_cell.value = "Open Photo" if english else "\u6253\u5f00\u7167\u7247"
                link_cell.hyperlink = image_path
                link_cell.style = "Hyperlink"

            for col_index in range(1, len(REVIEW_HEADERS) + 1):
                sheet.cell(row_index, col_index).alignment = Alignment(vertical="center", wrap_text=True)

        widths_by_header = {
            L_REVIEW: 12,
            L_PARSE_STATUS: 14,
            L_SOURCE: 26,
            L_OCR_TEXT: 48,
            L_LON: 16,
            L_LAT: 16,
            L_TIME: 22,
            L_CONFIDENCE: 30,
            L_PHOTO: 62 if image_mode != "none" else 20,
            L_PHOTO_LINK: 14,
            L_SOURCE_PATH: 56,
            L_TIPS: 40,
            L_MODEL_NOTE: 36,
            L_NOTE: 24,
        }
        from openpyxl.utils import get_column_letter

        for index, header in enumerate(REVIEW_HEADERS, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = widths_by_header.get(header, 18)

        sheet["A1"].comment = Comment(
            (
                "Auto Approved means the rules produced a high-confidence result. "
                "Only Needs Review rows require manual handling. Approved and Auto Approved rows can be written to photo metadata."
                if english
                else "\u81ea\u52a8\u901a\u8fc7\u8868\u793a\u89c4\u5219\u6216\u672c\u5730\u6a21\u578b\u5df2\u7ed9\u51fa\u9ad8\u7f6e\u4fe1\u7ed3\u679c\uff1b"
                "\u5f85\u6838\u884c\u624d\u9700\u4eba\u5de5\u5904\u7406\u3002\u901a\u8fc7\u548c\u81ea\u52a8\u901a\u8fc7\u90fd\u4f1a\u5199\u5165\u56fe\u7247\u5c5e\u6027\u3002"
            ),
            "Codex",
        )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        workbook.save(xlsx_path)


def read_review_rows(xlsx_path):
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("Missing dependency: openpyxl. Install it with: pip install openpyxl")

    workbook = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return []
    headers = [ZH_HEADER_BY_EN.get(clean_cell(value), clean_cell(value)) for value in rows[0]]
    result = []
    for row_index, row in enumerate(rows[1:], start=2):
        item = {"_row": row_index}
        for col_index, header in enumerate(headers):
            if header:
                value = row[col_index] if col_index < len(row) else None
                if header == L_REVIEW:
                    value = ZH_DECISION_BY_EN.get(clean_cell(value), value)
                elif header == L_PARSE_STATUS and clean_cell(value) == "EXIF Complete":
                    value = STATUS_EXIF_COMPLETE
                item[header] = value
        result.append(item)
    return result


def approved(value):
    return clean_cell(value).upper() in {v.upper() for v in APPROVED_VALUES}


def parse_review_coordinate(value, kind):
    text = normalize_text(value)
    if not text:
        raise ValueError("%s is empty" % kind)
    value = parse_coordinate_segment(text, kind)
    return abs(value)


def output_path_for_review_image(image_path, image_root, output_dir):
    try:
        return copy_output_path(image_path, image_root, output_dir)
    except Exception:
        return os.path.join(output_dir, os.path.basename(image_path))


def iter_source_files(image_root):
    for root, _, files in os.walk(image_root):
        for name in files:
            yield os.path.join(root, name)


def copy_all_source_files(image_root, output_dir, dry_run=False):
    copied = 0
    failed = 0
    for source_path in iter_source_files(image_root):
        dst_path = output_path_for_review_image(source_path, image_root, output_dir)
        try:
            if not dry_run:
                dst_dir = os.path.dirname(os.path.abspath(dst_path))
                if dst_dir and not os.path.isdir(dst_dir):
                    os.makedirs(dst_dir)
                if os.path.abspath(source_path) != os.path.abspath(dst_path):
                    shutil.copy2(source_path, dst_path)
            copied += 1
        except Exception as exc:
            failed += 1
            print("[COPY FAIL] %s -> %s" % (source_path, exc))
    return copied, failed


def read_exif_summary(image_path):
    result = {
        "exif_lat": "",
        "exif_lon": "",
        "GPSLatitudeRef": "",
        "GPSLongitudeRef": "",
        "DateTimeOriginal": "",
        "DateTimeDigitized": "",
        "ImageDateTime": "",
        "GPSMapDatum": "",
    }
    try:
        gps = read_gps_exif(image_path)
        if gps:
            result["exif_lat"] = "%.8f" % gps[0]
            result["exif_lon"] = "%.8f" % gps[1]
    except Exception:
        pass

    try:
        import piexif
        exif_dict = load_exif_dict(image_path)
        gps_ifd = exif_dict.get("GPS") or {}
        exif_ifd = exif_dict.get("Exif") or {}
        zeroth_ifd = exif_dict.get("0th") or {}

        def text_value(value):
            if value is None:
                return ""
            if isinstance(value, bytes):
                return value.decode("ascii", errors="replace")
            return str(value)

        result["GPSLatitudeRef"] = text_value(gps_ifd.get(piexif.GPSIFD.GPSLatitudeRef))
        result["GPSLongitudeRef"] = text_value(gps_ifd.get(piexif.GPSIFD.GPSLongitudeRef))
        result["GPSMapDatum"] = text_value(gps_ifd.get(piexif.GPSIFD.GPSMapDatum))
        result["DateTimeOriginal"] = text_value(exif_ifd.get(piexif.ExifIFD.DateTimeOriginal))
        result["DateTimeDigitized"] = text_value(exif_ifd.get(piexif.ExifIFD.DateTimeDigitized))
        result["ImageDateTime"] = text_value(zeroth_ifd.get(piexif.ImageIFD.DateTime))
    except Exception:
        pass

    # Pillow can read standard JPEG EXIF without piexif. This fallback keeps
    # analysis and review generation usable before the optional writer
    # dependency is installed.
    try:
        from PIL import Image as PILImage

        with PILImage.open(image_path) as image:
            exif = image.getexif()
            if not result["ImageDateTime"]:
                result["ImageDateTime"] = clean_cell(exif.get(306))
            if not result["DateTimeOriginal"]:
                result["DateTimeOriginal"] = clean_cell(exif.get(36867))
            if not result["DateTimeDigitized"]:
                result["DateTimeDigitized"] = clean_cell(exif.get(36868))

            if not result["exif_lat"] or not result["exif_lon"]:
                gps_ifd = exif.get_ifd(34853) if exif.get(34853) else {}
                lat_values = gps_ifd.get(2)
                lon_values = gps_ifd.get(4)
                lat_ref = gps_ifd.get(1)
                lon_ref = gps_ifd.get(3)

                def rational_value(value):
                    try:
                        return float(value)
                    except Exception:
                        return float(value[0]) / float(value[1])

                def gps_decimal(values, ref):
                    degree, minute, second = [rational_value(value) for value in values]
                    decimal = degree + minute / 60.0 + second / 3600.0
                    if clean_cell(ref).upper() in ("S", "W"):
                        decimal *= -1
                    return decimal

                if lat_values and lon_values and lat_ref and lon_ref:
                    result["exif_lat"] = "%.8f" % gps_decimal(lat_values, lat_ref)
                    result["exif_lon"] = "%.8f" % gps_decimal(lon_values, lon_ref)
                    result["GPSLatitudeRef"] = result["GPSLatitudeRef"] or clean_cell(lat_ref)
                    result["GPSLongitudeRef"] = result["GPSLongitudeRef"] or clean_cell(lon_ref)
    except Exception:
        pass

    return result


def result_report_image_path(row):
    output_path = clean_cell(row.get("output_image_path"))
    if output_path and not output_path.startswith("(") and os.path.isfile(output_path):
        return output_path
    input_path = clean_cell(row.get("input_image_path"))
    if input_path and os.path.isfile(input_path):
        return input_path
    return ""


def result_attribute_info(row):
    parts = [
        "\u5ba1\u6838\u7ed3\u679c\uff1a%s" % clean_cell(row.get("review_decision")),
        "\u7eac\u5ea6\uff1a%s" % clean_cell(row.get("review_lat_text")),
        "\u7ecf\u5ea6\uff1a%s" % clean_cell(row.get("review_lon_text")),
        "\u62cd\u6444\u65f6\u95f4\uff1a%s" % clean_cell(row.get("written_shooting_datetime")),
    ]
    note = clean_cell(row.get("review_note"))
    if note:
        parts.append("\u5ba1\u6838\u5907\u6ce8\uff1a%s" % note)
    return "\n".join(parts)


def write_result_report_xlsx(path, rows, max_width=520, max_height=340):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "\u5199\u5165\u7ed3\u679c"
    headers = [
        "\u72b6\u6001",
        "\u539f\u56e0",
        "\u89e3\u6790\u72b6\u6001",
        "\u8f93\u51fa\u7ed3\u679c\u56fe\u7247\u6587\u4ef6\u8def\u5f84",
        "\u8f93\u5165\u7684\u5c5e\u6027\u4fe1\u606f",
        "\u5bf9\u5e94\u7684\u56fe\u7247\u8def\u5f84",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    status_fills = {
        "OK": PatternFill("solid", fgColor="DFF0D8"),
        "SKIPPED": PatternFill("solid", fgColor="FCF8E3"),
        "FAIL": PatternFill("solid", fgColor="F2DEDE"),
    }

    for row_index, row in enumerate(rows, start=2):
        output_path = clean_cell(row.get("output_image_path"))
        sheet.append([
            row.get("status", ""),
            row.get("reason", ""),
            row.get("parse_status", ""),
            output_path,
            result_attribute_info(row),
            result_report_image_path(row),
        ])
        sheet.row_dimensions[row_index].height = 72
        fill = status_fills.get(row.get("status"))
        if fill:
            sheet.cell(row_index, 1).fill = fill

        image_path = result_report_image_path(row)
        if image_path:
            image_cell = sheet.cell(row_index, 5)
            image_cell.hyperlink = image_path
            image_cell.style = "Hyperlink"

        for cell in sheet[row_index]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for col, width in {
        "A": 12,
        "B": 34,
        "C": 18,
        "D": 64,
        "E": 36,
        "F": 64,
    }.items():
        sheet.column_dimensions[col].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)


def cleanup_old_generated_reports(output_dir):
    for name in ("write_result_report.csv", "image_exif_property_description.txt"):
        path = os.path.join(output_dir, name)
        try:
            if os.path.isfile(path):
                os.remove(path)
        except Exception as exc:
            print("[CLEANUP WARN] %s -> %s" % (path, exc))


def resolve_review_image_path(row):
    image_path = normalize_path(row.get(L_SOURCE_PATH))
    if image_path and os.path.isfile(image_path):
        return image_path
    return ""


def exif_write_plan(existing, has_latlon, has_time, overwrite_existing=False):
    """Return which reviewed fields may be written without replacing EXIF."""
    existing_has_gps = bool(
        clean_cell(existing.get("exif_lat"))
        and clean_cell(existing.get("exif_lon"))
    )
    existing_has_time = bool(
        clean_cell(existing.get("DateTimeOriginal"))
        or clean_cell(existing.get("DateTimeDigitized"))
        or clean_cell(existing.get("ImageDateTime"))
    )
    return {
        "existing_has_gps": existing_has_gps,
        "existing_has_time": existing_has_time,
        "write_latlon": bool(
            has_latlon and (overwrite_existing or not existing_has_gps)
        ),
        "write_time": bool(
            has_time and (overwrite_existing or not existing_has_time)
        ),
    }


def review_datetime_has_clock(value):
    """EXIF DateTimeOriginal requires a visible hour and minute."""
    text = clean_cell(value)
    return bool(
        re.search(
            r"(?:^|[T\s])(?:[01]?\d|2[0-3])[:：][0-5]\d(?:[:：][0-5]\d)?",
            text,
        )
    )


def review_datetime_for_exif(value):
    """Return a writable EXIF datetime and whether midnight was supplied."""
    text = clean_cell(value)
    if not text or review_datetime_has_clock(text):
        return text, False
    match = re.search(
        r"(?<!\d)((?:19|20)\d{2})[-:/\.]([01]?\d)[-:/\.]([0-3]?\d)(?!\d)",
        text,
    )
    if not match:
        return text, False
    year, month, day = [int(part) for part in match.groups()]
    try:
        date_value = datetime(year, month, day)
    except ValueError:
        return text, False
    return date_value.strftime("%Y:%m:%d 00:00:00"), True


def write_exif_from_review(args):
    image_root = os.path.abspath(args.image_root)
    output_dir = os.path.abspath(args.output_dir)
    # A dry run still writes its audit workbook, so the report directory must
    # exist even when no photo bytes are modified.
    if not args.in_place and not os.path.isdir(output_dir):
        os.makedirs(output_dir)

    copied_total = 0
    copy_failed = 0
    if not args.in_place:
        copied_total, copy_failed = copy_all_source_files(image_root, output_dir, args.dry_run)

    rows = read_review_rows(args.review_xlsx)
    reports = []
    total_ok = total_skip = total_fail = 0

    for row in rows:
        decision = clean_cell(row.get(L_REVIEW))
        parse_status = clean_cell(row.get(L_PARSE_STATUS))
        image_path = resolve_review_image_path(row)
        lat_text = clean_cell(row.get(L_LAT))
        lon_text = clean_cell(row.get(L_LON))
        report = {
            "status": "",
            "reason": "",
            "review_row": row.get("_row", ""),
            "review_decision": decision,
            "parse_status": parse_status,
            "input_image_path": image_path,
            "output_image_path": "",
            "written_lat": "",
            "written_lon": "",
            "review_lat_text": lat_text,
            "review_lon_text": lon_text,
            "written_shooting_datetime": "",
            "review_note": clean_cell(row.get(L_NOTE)),
        }

        try:
            if not approved(decision):
                report["status"] = "SKIPPED"
                report["reason"] = "review decision is not approved"
                if image_path:
                    report["output_image_path"] = output_path_for_review_image(image_path, image_root, output_dir) if not args.in_place else image_path
                total_skip += 1
                reports.append(report)
                continue

            if parse_status == STATUS_EXIF_COMPLETE:
                report["status"] = "SKIPPED"
                report["reason"] = "source image already has complete EXIF; not overwritten"
                if image_path:
                    report["output_image_path"] = output_path_for_review_image(image_path, image_root, output_dir) if not args.in_place else image_path
                total_skip += 1
                reports.append(report)
                continue

            if not image_path:
                report["status"] = "SKIPPED"
                report["reason"] = "image path not found in review row"
                total_skip += 1
                reports.append(report)
                continue
            image_ext = os.path.splitext(image_path)[1].lower()
            if image_ext not in WRITE_EXIF_EXTENSIONS:
                report["status"] = "SKIPPED"
                report["reason"] = "only JPG/JPEG/PNG can be written to EXIF"
                report["output_image_path"] = output_path_for_review_image(image_path, image_root, output_dir) if not args.in_place else image_path
                total_skip += 1
                reports.append(report)
                continue

            shooting_time = clean_cell(row.get(L_TIME))
            shooting_time_for_write, date_only_midnight = review_datetime_for_exif(
                shooting_time
            )
            has_date = bool(shooting_time)
            has_time = bool(
                shooting_time
                and (
                    review_datetime_has_clock(shooting_time)
                    or date_only_midnight
                )
            )
            has_latlon = bool(lat_text and lon_text)
            if not has_latlon and not has_date:
                report["status"] = "SKIPPED"
                if lat_text or lon_text:
                    report["reason"] = "latitude/longitude is incomplete and shooting datetime is empty"
                else:
                    report["reason"] = "latitude/longitude and shooting datetime are empty"
                report["output_image_path"] = output_path_for_review_image(image_path, image_root, output_dir) if not args.in_place else image_path
                total_skip += 1
                reports.append(report)
                continue

            lat = lon = None
            if has_latlon:
                lat = parse_review_coordinate(lat_text, "lat")
                lon = parse_review_coordinate(lon_text, "lon")
                validate_lat_lon(lat, lon)

            existing = read_exif_summary(image_path)
            overwrite_existing = bool(
                getattr(args, "overwrite_existing_exif", False)
            )
            plan = exif_write_plan(
                existing,
                has_latlon,
                has_time,
                overwrite_existing,
            )
            existing_has_gps = plan["existing_has_gps"]
            existing_has_time = plan["existing_has_time"]
            write_latlon = plan["write_latlon"]
            write_time = plan["write_time"]
            preserved = []
            warnings = []
            if date_only_midnight and write_time:
                warnings.append(
                    "date-only review value; wrote 00:00:00 as the clock time"
                )
            elif date_only_midnight:
                preserved.append(
                    "date-only review value (would use 00:00:00)"
                )
            if has_latlon and existing_has_gps and not overwrite_existing:
                preserved.append("existing GPS")
            if has_time and existing_has_time and not overwrite_existing:
                preserved.append("existing shooting datetime")

            if not write_latlon and not write_time:
                report["status"] = "SKIPPED"
                report["reason"] = (
                    "no missing EXIF field to write; preserved "
                    + ", ".join(preserved)
                )
                report["output_image_path"] = (
                    output_path_for_review_image(image_path, image_root, output_dir)
                    if not args.in_place
                    else image_path
                )
                total_skip += 1
                reports.append(report)
                continue

            if args.dry_run:
                output_path = "(dry-run, not written)"
            elif image_ext in JPG_EXTENSIONS and write_latlon:
                if args.in_place:
                    output_path = write_gps_exif(
                        image_path,
                        lat,
                        lon,
                        in_place=True,
                        datetime_original=shooting_time_for_write if write_time else None,
                    )
                else:
                    dst_path = output_path_for_review_image(image_path, image_root, output_dir)
                    output_path = write_gps_exif(
                        dst_path,
                        lat,
                        lon,
                        in_place=True,
                        datetime_original=shooting_time_for_write if write_time else None,
                    )
            elif image_ext in JPG_EXTENSIONS and write_time:
                if args.in_place:
                    output_path = write_datetime_exif(
                        image_path,
                        in_place=True,
                        datetime_original=shooting_time_for_write,
                    )
                else:
                    dst_path = output_path_for_review_image(image_path, image_root, output_dir)
                    output_path = write_datetime_exif(
                        dst_path,
                        in_place=True,
                        datetime_original=shooting_time_for_write,
                    )
            elif image_ext == ".png":
                if args.in_place:
                    output_path = write_png_exif(
                        image_path,
                        lat=lat if write_latlon else None,
                        lon=lon if write_latlon else None,
                        in_place=True,
                        datetime_original=shooting_time_for_write if write_time else None,
                    )
                else:
                    dst_path = output_path_for_review_image(image_path, image_root, output_dir)
                    output_path = write_png_exif(
                        dst_path,
                        lat=lat if write_latlon else None,
                        lon=lon if write_latlon else None,
                        in_place=True,
                        datetime_original=shooting_time_for_write if write_time else None,
                    )
            elif args.in_place:
                output_path = image_path
            else:
                output_path = output_path_for_review_image(image_path, image_root, output_dir)

            report["status"] = "OK"
            actions = []
            if write_latlon:
                actions.append("wrote GPS")
            if write_time:
                actions.append("wrote shooting datetime")
            if preserved:
                actions.append("preserved " + ", ".join(preserved))
            if warnings:
                actions.append("warning: " + ", ".join(warnings))
            report["reason"] = "; ".join(actions)
            report["output_image_path"] = output_path
            if write_latlon:
                report["written_lat"] = "%.8f" % lat
                report["written_lon"] = "%.8f" % lon
            if write_time:
                report["written_shooting_datetime"] = shooting_time_for_write
            if output_path and not output_path.startswith("("):
                report.update(read_exif_summary(output_path))
            total_ok += 1
        except Exception as exc:
            report["status"] = "FAIL"
            report["reason"] = str(exc)
            total_fail += 1

        reports.append(report)

    report_xlsx = os.path.join(output_dir, "write_result_report.xlsx")
    write_result_report_xlsx(report_xlsx, reports, args.photo_display_max_width, args.photo_display_max_height)
    cleanup_old_generated_reports(output_dir)

    print("Done. OK=%d SKIPPED=%d FAIL=%d" % (total_ok, total_skip, total_fail))
    if not args.in_place:
        print("Copied source files: %d copy_failed=%d" % (copied_total, copy_failed))
    print("Result report XLSX: %s" % report_xlsx)
    return 0 if total_fail == 0 else 2
